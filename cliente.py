import pandas as pd
from pathlib import Path
import streamlit as st
import logging
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
class ClienteManager:
    def __init__(self, session_state):
        self.session_state = session_state
        self.df_clientes = pd.DataFrame()
        self.df_tarifas = pd.DataFrame()
        self.tarifas_arconel = []
        self.tarifa_info = {}
        self.current_client_index = 0
        self.nuevo_cliente = True

        self.campos_principales = [
            "Empresa", "Persona de Contacto", "Dirección", "Ciudad", "Provincia",
            "Velocidad de Viento", "Tipo de Tarifa ARCONEL", "Tarifario",
            "Tensión eléctrica", "Voltaje de entrada", "Factor de Multiplicación",
            "Factor de Corrección", "Factor de Potencia", "Arrays", "Días Facturados"
        ]

        self.campos_consumo = [
            "kW/h A", "Total A", "kW/h B", "Total B",
            "kW/h C", "Total C", "kW/h D", "Total D",
            "kW/h Demanda", "Total Demanda"
        ]

        self.inicializar_session_state()
        self.cargar_datos()

    def inicializar_session_state(self):
        if 'cliente_data' not in self.session_state:
            self.session_state['cliente_data'] = {campo: "" for campo in self.campos_principales}
        
        if 'consumo_data' not in self.session_state:
            self.session_state['consumo_data'] = {campo: "0" for campo in self.campos_consumo}
        
        if 'array_data' not in self.session_state:
            self.session_state['array_data'] = {f"Array_{i}_{dim}": "0" for i in range(1, 9) for dim in ['X', 'Y']}
        
        if 'requerimiento_data' not in self.session_state:
            self.session_state['requerimiento_data'] = {
                "Demanda (08h00-22h00)_consumo": "0.00",
                "Demanda (08h00-22h00)_requerido": "0.00",
                "Demanda (08h00-22h00)_ideal": "0.00",
                "Demanda (24 Horas)_consumo": "0.00",
                "Demanda (24 Horas)_requerido": "0.00",
                "Demanda (24 Horas)_ideal": "0.00",
                "Respaldo (24H)_consumo": "0.00",
                "Respaldo (24H)_requerido": "0.00",
                "Respaldo (24H)_ideal": "0.00",
                "Respaldo (4 Horas de Respaldo)_consumo": "0.00",
                "Respaldo (4 Horas de Respaldo)_requerido": "0.00",
                "Respaldo (4 Horas de Respaldo)_ideal": "0.00"
            }

        if 'equipamiento_seleccionado' not in self.session_state:
            self.session_state['equipamiento_seleccionado'] = {
                "Paneles Solares": "",
                "Inversores": "",
                "Baterías": "",
                "Convertidor de Alto Voltaje DC": "",
                "Materiales Estructura Solar": "",
                "Materiales DB": "",
                "Panel Solar": {},
                "Inversor": {}
            }

        if 'distribucion_data' not in self.session_state:
            self.session_state['distribucion_data'] = {
                'panel_seleccionado': None,
                'paneles_por_array': {},
                'total_paneles': 0,
                'total_potencia': 0.0,
                'distribucion_actualizada': False,
                'arrays_config': [],
                'layout_data': {
                    'orientacion': 'Portrait',
                    'azimuth': '152°',
                    'pitch': '5°',
                    'ground_clearence': '',
                    'espacio_bordes': '1,00 Metros',
                    'filas_x_columna': '',
                    'columnas': '',
                    'total_filas_array': '',
                    'paneles_fila': '',
                    'paneles_columna': '',
                    'paneles_array': '',
                    'largo_fila': '',
                    'separacion_columnas': '',
                    'separacion_filas': ''
                }
            }

    # CORRECCIÓN CLAVE: MÉTODO DE GUARDADO FUNCIONAL
    
    def guardar_proyecto_en_excel(self):
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows

            desktop_path = Path.home() / 'Desktop'
            folder_path = desktop_path / 'Calculadora_Solar'
            folder_path.mkdir(parents=True, exist_ok=True)
            archivo = folder_path / "Calculadora Solar.xlsx"

            empresa = self.session_state.get("cliente_data", {}).get("Empresa", "").strip()
            if not empresa:
                st.error("❌ El campo 'Empresa' es obligatorio.")
                return

            def convert_dict(d): return {k: str(v) for k, v in d.items()} if isinstance(d, dict) else {}

            # Armar cada bloque como DataFrame
            df_clientes = pd.DataFrame([{
                **self.session_state.get("cliente_data", {}),
                **self.session_state.get("consumo_data", {}),
                **self.session_state.get("requerimiento_data", {}),
                **{
                    f"Array {i} {dim}": self.session_state["array_data"].get(f"Array_{i}_{dim}", "0")
                    for i in range(1, 9) for dim in ['X', 'Y']
                }
            }])

            equip = self.session_state.get("equipamiento_seleccionado", {})
            df_equipamiento = pd.DataFrame([{
                "Empresa": empresa,
                "Paneles Solares": equip.get("Paneles Solares", ""),
                "Inversores": equip.get("Inversores", ""),
                "Baterías": equip.get("Baterías", ""),
                "Convertidor de Alto Voltaje DC": equip.get("Convertidor de Alto Voltaje DC", ""),
                "Materiales Estructura Solar": equip.get("Materiales Estructura Solar", ""),
                "Materiales DB": equip.get("Materiales DB", "")
            }])

            df_distribucion = pd.DataFrame(self.session_state.get("distribucion_data", {}).get("arrays_config", []))
            if not df_distribucion.empty:
                df_distribucion["Empresa"] = empresa

            estructura_raw = self.session_state.get("estructura_total_materiales", {})
            df_estructura = pd.DataFrame([{
                "Empresa": empresa,
                "Materiales": str(estructura_raw)
            }])

            df_generacion = pd.DataFrame([convert_dict(self.session_state.get("generacion_data", {}))])
            df_generacion["Empresa"] = empresa

            df_componentes = pd.DataFrame([convert_dict(self.session_state.get("componentes_principales", {}))])
            df_componentes["Empresa"] = empresa

            df_costos = pd.DataFrame(self.session_state.get("costos_data", []))
            if not df_costos.empty:
                df_costos["Empresa"] = empresa

            df_analisis = pd.DataFrame([convert_dict(self.session_state.get("analisis_economico", {}))])
            df_analisis["Empresa"] = empresa

            hojas_proyecto = {
                "Clientes": df_clientes,
                "Equipamiento Seleccionado": df_equipamiento,
                "Distribución Solar": df_distribucion,
                "Estructura Solar": df_estructura,
                "Generación": df_generacion,
                "Componentes Principales": df_componentes,
                "Costos": df_costos,
                "Análisis Económico": df_analisis
            }

            hojas_catalogo = [
                "Pliego Tarifario", "Paneles Solares", "Inversores", "Baterías",
                "Convertidor de Alto Voltaje DC", "Materiales Estructura Solar", "Materiales DB"
            ]

            # Cargar archivo si existe
            if archivo.exists():
                wb = load_workbook(archivo)
            else:
                from openpyxl import Workbook
                wb = Workbook()
                wb.remove(wb.active)

            for hoja, df_nuevo in hojas_proyecto.items():
                if hoja in hojas_catalogo:
                    continue

                # Reemplazar cliente en hoja existente
                if hoja in wb.sheetnames:
                    hoja_df = pd.read_excel(archivo, sheet_name=hoja)
                    if "Empresa" in hoja_df.columns:
                        hoja_df = hoja_df[hoja_df["Empresa"] != empresa]
                        df_final = pd.concat([hoja_df, df_nuevo], ignore_index=True)
                    else:
                        df_final = df_nuevo
                    wb.remove(wb[hoja])
                else:
                    df_final = df_nuevo

                ws = wb.create_sheet(hoja)
                for row in dataframe_to_rows(df_final, index=False, header=True):
                    ws.append([str(c) if isinstance(c, dict) else c for c in row])

            wb.save(archivo)
            st.success(f"✅ Proyecto '{empresa}' guardado correctamente.")
            self.session_state['nuevo_cliente'] = False

        except Exception as e:
            st.error(f"❌ Error al guardar el proyecto: {str(e)}")
            import traceback
            st.error(traceback.format_exc())


    def cargar_proyecto_completo(self, nombre_cliente):
        try:
            desktop_path = Path.home() / 'Desktop'
            folder_path = desktop_path / 'Calculadora_Solar'
            archivo = folder_path / "Calculadora Solar.xlsx"
        
            if not archivo.exists():
                st.error(f"Archivo no encontrado: {archivo}")
                return

            # Leer todas las hojas
            datos = pd.read_excel(archivo, sheet_name=None)
        
            # Cargar datos básicos del cliente
            cliente_df = datos.get("Clientes", pd.DataFrame())
            cliente_row = cliente_df[cliente_df["Empresa"] == nombre_cliente]
            if cliente_row.empty:
                st.error(f"Cliente '{nombre_cliente}' no encontrado")
                return
        
            # ✅ Cargar datos desde hoja "Clientes"
            self.cargar_datos_cliente(cliente_row.iloc[0])

            # ✅ Cargar equipamiento
            equipamiento_df = datos.get("Equipamiento Seleccionado", pd.DataFrame())
            if not equipamiento_df.empty:
                eq_row = equipamiento_df[equipamiento_df["Empresa"] == nombre_cliente]
                if not eq_row.empty:
                    eq_data = eq_row.iloc[0].fillna("").to_dict()
                    self.session_state["equipamiento_seleccionado"] = {
                        "Paneles Solares": eq_data.get("Paneles Solares", ""),
                        "Inversores": eq_data.get("Inversores", ""),
                        "Baterías": eq_data.get("Baterías", ""),
                        "Convertidor de Alto Voltaje DC": eq_data.get("Convertidor de Alto Voltaje DC", ""),
                        "Materiales Estructura Solar": eq_data.get("Materiales Estructura Solar", ""),
                        "Materiales DB": eq_data.get("Materiales DB", ""),
                        "Panel Solar": {},
                        "Inversor": {}
                    }
                    self._cargar_datos_tecnicos_desde_catalogo()

            # Cargar distribución solar
            distribucion_df = datos.get("Distribución Solar", pd.DataFrame())
            if not distribucion_df.empty:
                dist_row = distribucion_df[distribucion_df["Empresa"] == nombre_cliente]
                if not dist_row.empty:
                    self.session_state["distribucion_data"]["arrays_config"] = dist_row.replace({np.nan: None}).to_dict(orient='records')

            # Cargar estructura solar
            estructura_df = datos.get("Estructura Solar", pd.DataFrame())
            if not estructura_df.empty:
                est_row = estructura_df[estructura_df["Empresa"] == nombre_cliente]
                if not est_row.empty:
                    self.session_state["estructura_total_materiales"] = est_row.replace({np.nan: None}).to_dict(orient='records')

            # Cargar generación
            generacion_df = datos.get("Generación", pd.DataFrame())
            if not generacion_df.empty:
                gen_row = generacion_df[generacion_df["Empresa"] == nombre_cliente]
                if not gen_row.empty:
                    self.session_state["generacion_data"] = gen_row.iloc[0].fillna("0").to_dict()

            # Cargar componentes principales
            componentes_df = datos.get("Componentes Principales", pd.DataFrame())
            if not componentes_df.empty:
                comp_row = componentes_df[componentes_df["Empresa"] == nombre_cliente]
                if not comp_row.empty:
                    self.session_state["componentes_principales"] = comp_row.iloc[0].fillna("0").to_dict()

            # Cargar costos
            costos_df = datos.get("Costos", pd.DataFrame())
            if not costos_df.empty:
                cost_row = costos_df[costos_df["Empresa"] == nombre_cliente]
                if not cost_row.empty:
                    self.session_state["costos_data"] = cost_row.replace({np.nan: None}).to_dict(orient='records')

            # Cargar análisis económico
            analisis_df = datos.get("Análisis Económico", pd.DataFrame())
            if not analisis_df.empty:
                ana_row = analisis_df[analisis_df["Empresa"] == nombre_cliente]
                if not ana_row.empty:
                    self.session_state["analisis_economico"] = ana_row.iloc[0].fillna("0").to_dict()

            st.success(f"✅ Proyecto '{nombre_cliente}' cargado exitosamente")
            self.nuevo_cliente = False

            # Verificar carga
            self._verificar_carga_correcta()

        except Exception as e:
            st.error(f"❌ Error al cargar proyecto: {str(e)}")
            import traceback
            st.error(traceback.format_exc())

            
    def cargar_datos_cliente(self, fila_cliente):
        """Carga los datos básicos del cliente desde una fila del DataFrame"""
        st.subheader("📥 Cargando datos del cliente desde Excel...")

        # Normalizar nombres de columnas si es necesario
        def normalizar_clave(clave):
            clave = clave.lower().strip()
            reemplazos = {
                "kwh a": "kW/h A", "total a": "Total A",
                "kwh b": "kW/h B", "total b": "Total B",
                "kwh c": "kW/h C", "total c": "Total C",
                "kwh d": "kW/h D", "total d": "Total D",
                "kwh demanda": "kW/h Demanda", "total demanda": "Total Demanda"
            }
            return reemplazos.get(clave, clave)

        # Cargar campos principales
        for campo in self.campos_principales:
            valor = fila_cliente.get(campo, "")
            if campo == "Arrays":
                try:
                    valor = str(int(float(valor))) if valor and str(valor).strip() else "0"
                except:
                    valor = "0"
            self.session_state['cliente_data'][campo] = "" if pd.isna(valor) else str(valor)

        # Cargar campos de consumo energético
        for campo in self.campos_consumo:
            valor = fila_cliente.get(campo, "")
            st.write(f"🟨 {campo} => {valor}")
            self.session_state['consumo_data'][campo] = "0" if pd.isna(valor) else str(valor)
            

        # Cargar datos de arrays físicos
        for i in range(1, 9):
            for dim in ['X', 'Y']:
                col_standard = f"Array {i} {dim}"
                valor = fila_cliente.get(col_standard, "")
                self.session_state['array_data'][f"Array_{i}_{dim}"] = "0" if pd.isna(valor) else str(valor)

        # Cargar requerimientos energéticos
        requerimientos = [
            "Demanda (08h00-22h00)",
            "Demanda (24 Horas)",
            "Respaldo (24H)",
            "Respaldo (4 Horas de Respaldo)"
        ]

        for req in requerimientos:
            for sufijo in ['consumo', 'requerido', 'ideal']:
                key = f"{req}_{sufijo}"
                if key in fila_cliente:
                    self.session_state['requerimiento_data'][key] = "0.00" if pd.isna(fila_cliente[key]) else str(fila_cliente[key])

        st.write("🔍 kWh A leído desde Excel:", fila_cliente.get("kW/h A"))
        st.write("🔁 Valor guardado en session_state:", self.session_state["consumo_data"].get("kW/h A"))


    def _cargar_datos_tecnicos_desde_catalogo(self):
        """Carga los datos técnicos del panel e inversor desde el catálogo"""
        try:
            panel_nombre = self.session_state["equipamiento_seleccionado"].get("Paneles Solares", "")
            inversor_nombre = self.session_state["equipamiento_seleccionado"].get("Inversores", "")

            if not panel_nombre and not inversor_nombre:
                return

            desktop_path = Path.home() / 'Desktop'
            folder_path = desktop_path / 'Calculadora_Solar'
            path_catalogo = folder_path / 'Calculadora Solar.xlsx'
            if not path_catalogo.exists():
                st.warning("⚠️ Archivo de catálogo técnico no encontrado")
                return

            catalogos = pd.read_excel(path_catalogo, sheet_name=None)

            # Cargar datos del panel solar
            if panel_nombre:
                paneles_df = catalogos.get("Paneles Solares", pd.DataFrame())
                if not paneles_df.empty:
                    panel_match = paneles_df[
                        paneles_df["Modelo"].str.strip().str.lower() == panel_nombre.strip().lower()
                    ]
                    if not panel_match.empty:
                        self.session_state["equipamiento_seleccionado"]["Panel Solar"] = panel_match.iloc[0].fillna("").to_dict()

            # Cargar datos del inversor
            if inversor_nombre:
                inversores_df = catalogos.get("Inversores", pd.DataFrame())
                if not inversores_df.empty:
                    inversor_match = inversores_df[
                        inversores_df["Modelo"].str.strip().str.lower() == inversor_nombre.strip().lower()
                    ]
                    if not inversor_match.empty:
                        self.session_state["equipamiento_seleccionado"]["Inversor"] = inversor_match.iloc[0].fillna("").to_dict()

        except Exception as e:
            st.warning(f"⚠️ Error al cargar datos técnicos: {str(e)}")

    def _verificar_carga_correcta(self):
        """Verifica que todos los datos se cargaron correctamente"""
        empresa = self.session_state['cliente_data'].get("Empresa", "SIN NOMBRE")
        st.write(f"### Verificación de carga para: {empresa}")
        
        # Verificar equipamiento
        equipamiento = self.session_state.get("equipamiento_seleccionado", {})
        panel = equipamiento.get("Paneles Solares", "NO CARGADO")
        inversor = equipamiento.get("Inversores", "NO CARGADO")
        
        st.write(f"- **Panel solar seleccionado:** {panel}")
        st.write(f"- **Inversor seleccionado:** {inversor}")
        
        if equipamiento.get("Panel Solar"):
            st.success("✅ Datos técnicos del panel cargados")
        else:
            st.warning("⚠️ Datos técnicos del panel NO cargados")
            
        if equipamiento.get("Inversor"):
            st.success("✅ Datos técnicos del inversor cargados")
        else:
            st.warning("⚠️ Datos técnicos del inversor NO cargados")
        
        # Verificar otros componentes
        st.write(f"- **Configuración de arrays:** {len(self.session_state.get('distribucion_data', {}).get('arrays_config', []))} arrays")
        st.write(f"- **Materiales de estructura:** {len(self.session_state.get('estructura_total_materiales', []))} items")
        st.write(f"- **Datos de generación:** {'Sí' if self.session_state.get('generacion_data') else 'No'}")
        st.write(f"- **Componentes principales:** {'Sí' if self.session_state.get('componentes_principales') else 'No'}")
        st.write(f"- **Análisis económico:** {'Sí' if self.session_state.get('analisis_economico') else 'No'}")

    def mostrar_equipamiento_seleccionado(self):
        """Muestra los selectores de equipos con los valores cargados"""
        # Obtener listas disponibles de paneles e inversores
        lista_paneles = self._obtener_lista_paneles()
        lista_inversores = self._obtener_lista_inversores()
    
        # Obtener valores actualmente seleccionados
        panel_actual = self.session_state["equipamiento_seleccionado"].get("Paneles Solares", "")
        inversor_actual = self.session_state["equipamiento_seleccionado"].get("Inversores", "")
    
        # Mostrar selector de paneles con el valor cargado
        col1, col2 = st.columns(2)
        with col1:
            nuevo_panel = st.selectbox(
                "Panel Solar",
                options=lista_paneles,
                index=lista_paneles.index(panel_actual) if panel_actual in lista_paneles else 0,
                key="select_panel"
            )
        
            # Actualizar si cambió la selección
            if nuevo_panel != panel_actual:
                self.session_state["equipamiento_seleccionado"]["Paneles Solares"] = nuevo_panel
                self._cargar_datos_tecnicos_desde_catalogo()
                st.rerun()
    
        with col2:
            nuevo_inversor = st.selectbox(
                "Inversor",
                options=lista_inversores,
                index=lista_inversores.index(inversor_actual) if inversor_actual in lista_inversores else 0,
                key="select_inversor"
            )
        
            # Actualizar si cambió la selección
            if nuevo_inversor != inversor_actual:
                self.session_state["equipamiento_seleccionado"]["Inversores"] = nuevo_inversor
                self._cargar_datos_tecnicos_desde_catalogo()
                st.rerun()
    
        # Mostrar datos técnicos si están disponibles
        if self.session_state["equipamiento_seleccionado"].get("Panel Solar"):
            st.subheader("Especificaciones del Panel Solar")
            datos_panel = self.session_state["equipamiento_seleccionado"]["Panel Solar"]
            st.json(datos_panel)
    
        if self.session_state["equipamiento_seleccionado"].get("Inversor"):
            st.subheader("Especificaciones del Inversor")
            datos_inversor = self.session_state["equipamiento_seleccionado"]["Inversor"]
            st.json(datos_inversor)

    def _obtener_lista_paneles(self):
        """Devuelve lista de paneles disponibles (debe implementarse según tu catálogo)"""
        try:
            desktop_path = Path.home() / 'Desktop'
            folder_path = desktop_path / 'Calculadora_Solar'
            path_catalogo = folder_path / 'Calculadora Solar.xlsx'
            if path_catalogo.exists():
                paneles_df = pd.read_excel(path_catalogo, sheet_name="Paneles Solares")
                return paneles_df["Modelo"].tolist()
        except:
            pass
        return ["Jinko Tiger Neo Mono 605", "Jinko Tiger Neo Mono 595"]  # Default

    def _obtener_lista_inversores(self):
        """Devuelve lista de inversores disponibles (debe implementarse según tu catálogo)"""
        try:
            desktop_path = Path.home() / 'Desktop'
            folder_path = desktop_path / 'Calculadora_Solar'
            path_catalogo = folder_path / 'Calculadora Solar.xlsx'
            if path_catalogo.exists():
                inversores_df = pd.read_excel(path_catalogo, sheet_name="Inversores")
                return inversores_df["Modelo"].tolist()
        except:
            pass
        return ["Growatt MAC 70KTL3-X MV", "Deye SUN-50K-SG01HP3-US-BM4-277V"]  # Default

    def redondear_especial(self, valor):
        try:
            valor = float(str(valor).replace(",", "."))
            entero = int(valor)
            decimal = valor - entero
            if decimal >= 0.5:
                return entero + 1
            return entero
        except:
            return 0
    
    def cargar_datos(self):
        try:
            desktop_path = Path.home() / 'Desktop'
            folder_path = desktop_path / 'Calculadora_Solar'
            folder_path.mkdir(parents=True, exist_ok=True)
            
            # Cargar clientes
            clientes_path = folder_path / 'Calculadora Solar.xlsx'
            if clientes_path.exists():
                self.df_clientes = pd.read_excel(clientes_path, sheet_name=0)
                if not self.df_clientes.empty:
                    self.df_clientes.columns = [col.strip() for col in self.df_clientes.columns]
                    
                    # Asegurar que todos los campos existan
                    for campo in self.campos_principales + self.campos_consumo:
                        if campo not in self.df_clientes.columns:
                            self.df_clientes[campo] = ""
                    
                    # Normalizar nombres de columnas de arrays
                    for i in range(1, 9):
                        for dim in ['X', 'Y']:
                            col_standard = f"Array {i} {dim}"
                            col_alternate = f"Array{i}{dim}"
                            
                            if col_standard in self.df_clientes.columns:
                                continue
                            elif col_alternate in self.df_clientes.columns:
                                self.df_clientes.rename(columns={col_alternate: col_standard}, inplace=True)
                            else:
                                self.df_clientes[col_standard] = ""
                    
                    # Normalizar campo Arrays
                    if 'Arrays' in self.df_clientes.columns:
                        self.df_clientes['Arrays'] = pd.to_numeric(
                            self.df_clientes['Arrays'], 
                            errors='coerce'
                        ).fillna(0).astype(int)
                    
                    # Reemplazar NaN con cadenas vacías
                    self.df_clientes = self.df_clientes.fillna("")
            
            # Cargar tarifas desde la hoja "Pliego Tarifario"
            tarifas_path = folder_path / 'Calculadora Solar.xlsx'
            if tarifas_path.exists():
                try:
                    self.df_tarifas = pd.read_excel(tarifas_path, sheet_name="Pliego Tarifario")
                    
                    if not self.df_tarifas.empty:
                        self.df_tarifas.columns = [col.strip() for col in self.df_tarifas.columns]
                        
                        # Buscar columnas por nombres esperados
                        codigo_col = next((col for col in self.df_tarifas.columns if 'código' in col.lower() or 'codigo' in col.lower()), None)
                        tipo_col = next((col for col in self.df_tarifas.columns if 'tipo' in col.lower() or 'descripción' in col.lower() or 'descripcion' in col.lower()), None)
                        tarifario_col = next((col for col in self.df_tarifas.columns if 'tarifario' in col.lower()), None)
                        tension_col = next((col for col in self.df_tarifas.columns if 'tensión' in col.lower() or 'tension' in col.lower()), None)
                        voltaje_col = next((col for col in self.df_tarifas.columns if 'voltaje' in col.lower()), None)
                        
                        if codigo_col:
                            for _, row in self.df_tarifas.iterrows():
                                codigo = str(row[codigo_col]) if pd.notna(row[codigo_col]) else ""
                                if codigo:
                                    tipo = str(row[tipo_col]) if tipo_col and pd.notna(row[tipo_col]) else ""
                                    tarifario = str(row[tarifario_col]) if tarifario_col and pd.notna(row[tarifario_col]) else ""
                                    tension = str(row[tension_col]) if tension_col and pd.notna(row[tension_col]) else ""
                                    voltaje = str(row[voltaje_col]) if voltaje_col and pd.notna(row[voltaje_col]) else ""
                                    
                                    self.tarifa_info[codigo] = {
                                        'tipo': tipo,
                                        'horarios': self.determinar_si_tiene_horarios(codigo, tipo),
                                        'tarifario': tarifario,
                                        'tension': tension,
                                        'voltaje': voltaje
                                    }
                            self.tarifas_arconel = list(self.tarifa_info.keys())
                except Exception as e:
                    st.error(f"Error al cargar tarifas: {str(e)}")
                    self.tarifas_arconel = []

        except Exception as e:
            st.error(f"Error al cargar datos: {str(e)}")
    
    def determinar_si_tiene_horarios(self, codigo, tipo):
        if codigo in ['BTCRSD01-BT', 'BTCGCD30-BT', 'MTCGCD32-MT']:
            return True
        return bool(tipo and ('horario' in tipo.lower() or 'pico' in tipo.lower()))
    
    def formatear_decimal(self, valor, decimales=2):
        try:
            valor = float(valor)
            return f"{valor:.{decimales}f}"
        except:
            return f"0.{'0'*decimales}"
    
    def mostrar_pestana(self):
        st.header("📝 Datos del Cliente")
        
        # Botón de búsqueda en la parte superior
        if st.button("🔍 Buscar Cliente"):
            self.session_state['mostrar_busqueda'] = True
            
        if self.session_state.get('mostrar_busqueda', False):
            self.mostrar_busqueda_clientes()
            return
            
        # Mostrar el formulario completo
        self.mostrar_formulario_cliente()
        
        
    def mostrar_formulario_cliente(self):
        for campo in self.campos_principales:
            if campo == "Tipo de Tarifa ARCONEL":
                current_value = self.session_state['cliente_data'].get(campo, "")
                
                # Verificar si hay opciones disponibles
                if not self.tarifas_arconel:
                    st.warning("No se cargaron tarifas desde el archivo Excel")
                    selected = st.selectbox(campo, options=[], index=0)
                else:
                    # Selección con índice seguro
                    index = self.tarifas_arconel.index(current_value) if current_value in self.tarifas_arconel else 0
                    selected = st.selectbox(campo, options=self.tarifas_arconel, index=index, key="select_tarifa")
                    
                if selected != current_value:
                    self.session_state['cliente_data'][campo] = selected
                    self.actualizar_campos_tarifa(selected)
            elif campo == "Arrays":
                current_value = self.session_state['cliente_data'].get(campo, "0")
                try:
                    current_value = int(float(current_value)) if current_value else 0
                except:
                    current_value = 0
                
                new_value = st.text_input(
                    campo,
                    value=str(current_value),
                    key=f"input_{campo}"
                )
                
                try:
                    int_value = int(float(new_value)) if new_value else 0
                    self.session_state['cliente_data'][campo] = str(int_value)
                except:
                    self.session_state['cliente_data'][campo] = "0"
            else:
                self.session_state['cliente_data'][campo] = st.text_input(
                    campo,
                    value=self.session_state['cliente_data'].get(campo, ""),
                    key=f"input_{campo}"
                )
        
        self.mostrar_consumo_energetico()
        self.mostrar_espacio_fisico()
        self.mostrar_requerimiento_energetico()
    
    def mostrar_consumo_energetico(self):
        st.header("Consumo Energético")

        st.markdown("""
        <style>
        .consumo-table th {
            font-size: 12px;
            font-weight: bold;
            padding: 5px;
            text-align: center;
        }
        .consumo-table td {
            font-size: 12px;
            padding: 5px;
        }
        </style>
        """, unsafe_allow_html=True)

        tipos = ['A', 'B', 'C', 'D']
        total_kwh = 0.0
        total_monto = 0.0
        clave_cliente = self.session_state['cliente_data'].get("Empresa", "nuevo")

        cols = st.columns([2, 2, 2, 2])
        headers = ["**Tipo**", "**Costo/kWh**", "**kWh**", "**Total ($)**"]
        for i, header in enumerate(headers):
            cols[i].markdown(header, unsafe_allow_html=True)

        for tipo in tipos:
            cols = st.columns([2, 2, 2, 2])

            cols[0].write(f"Consumo {tipo}")

            kwh_key = f"kW/h {tipo}"
            total_key = f"Total {tipo}"

            kwh = self.session_state['consumo_data'].get(kwh_key, "0")
            total = self.session_state['consumo_data'].get(total_key, "0")

            try:
                kwh_val = float(kwh.replace(",", ".")) if kwh and kwh.replace(",", "").replace(".", "").isdigit() else 0.0
                total_val = float(total.replace(",", ".")) if total and total.replace(",", "").replace(".", "").isdigit() else 0.0
                costo = total_val / kwh_val if kwh_val > 0 else 0.0
                cols[1].write(f"${self.formatear_decimal(costo)}")

                total_kwh += kwh_val
                total_monto += total_val
            except:
                cols[1].write("")

            new_kwh = cols[2].text_input(
                f"kWh {tipo}", 
                value=kwh,
                key=f"kwh_input_{tipo}_{clave_cliente}",
                label_visibility="collapsed"
            )
            self.session_state['consumo_data'][kwh_key] = new_kwh if new_kwh else "0"

            new_total = cols[3].text_input(
                f"Total {tipo}", 
                value=total,
                key=f"total_input_{tipo}_{clave_cliente}",
                label_visibility="collapsed"
            )
            self.session_state['consumo_data'][total_key] = new_total if new_total else "0"

        # DEMANDA FACTURABLE
        st.subheader("Demanda Facturable")
        cols = st.columns([2, 2, 2, 2])
        cols[0].write("Demanda Facturable")

        kwh_key = "kW/h Demanda"
        total_key = "Total Demanda"

        kwh = self.session_state['consumo_data'].get(kwh_key, "0")
        total = self.session_state['consumo_data'].get(total_key, "0")

        try:
            kwh_val = float(kwh.replace(",", ".")) if kwh and kwh.replace(",", "").replace(".", "").isdigit() else 0.0
            total_val = float(total.replace(",", ".")) if total and total.replace(",", "").replace(".", "").isdigit() else 0.0
            costo = total_val / kwh_val if kwh_val > 0 else 0.0
            cols[1].write(f"${self.formatear_decimal(costo)}")

            total_kwh += kwh_val
            total_monto += total_val
        except:
            cols[1].write("")

        new_kwh = cols[2].text_input(
            "kWh Demanda",
            value=kwh,
            key=f"kwh_input_demanda_{clave_cliente}",
            label_visibility="collapsed"
        )
        self.session_state['consumo_data'][kwh_key] = new_kwh if new_kwh else "0"

        new_total = cols[3].text_input(
            "Total Demanda",
            value=total,
            key=f"total_input_demanda_{clave_cliente}",
            label_visibility="collapsed"
        )
        self.session_state['consumo_data'][total_key] = new_total if new_total else "0"

        # TOTALES
        st.subheader("Totales")
        col1, col2 = st.columns(2)

        # Solo A, B, C, D para el consumo total
        consumo_total_kwh = 0.0
        for tipo in ['A', 'B', 'C', 'D']:
            kwh_val = self.session_state['consumo_data'].get(f"kW/h {tipo}", "0")
            try:
                consumo_total_kwh += float(kwh_val.replace(",", ".")) if kwh_val.replace(",", "").replace(".", "").isdigit() else 0.0
            except:
                pass

        col1.metric("Consumo Total (kWh)", self.formatear_decimal(consumo_total_kwh))
        col2.metric("Monto Total ($)", f"${self.formatear_decimal(total_monto)}")

        if consumo_total_kwh > 0:
            st.write(f"**Costo promedio:** ${self.formatear_decimal(total_monto / consumo_total_kwh)}/kWh")
    
    def calcular_requerimientos(self):
        try:
            consumo_a = float(self.session_state['consumo_data'].get("kW/h A", "0").replace(",", ".")) if self.session_state['consumo_data'].get("kW/h A", "0").replace(",", "").replace(".", "").isdigit() else 0.0
            consumo_b = float(self.session_state['consumo_data'].get("kW/h B", "0").replace(",", ".")) if self.session_state['consumo_data'].get("kW/h B", "0").replace(",", "").replace(".", "").isdigit() else 0.0
            consumo_c = float(self.session_state['consumo_data'].get("kW/h C", "0").replace(",", ".")) if self.session_state['consumo_data'].get("kW/h C", "0").replace(",", "").replace(".", "").isdigit() else 0.0
            consumo_d = float(self.session_state['consumo_data'].get("kW/h D", "0").replace(",", ".")) if self.session_state['consumo_data'].get("kW/h D", "0").replace(",", "").replace(".", "").isdigit() else 0.0
            
            dias_facturados_str = self.session_state['cliente_data'].get("Días Facturados", "30")
            try:
                dias_facturados = float(dias_facturados_str.replace(",", ".")) if dias_facturados_str.replace(",", "").replace(".", "").isdigit() else 30.0
            except:
                dias_facturados = 30.0
            
            dias_facturados = max(dias_facturados, 1)
            
            consumo_diario_a = consumo_a / dias_facturados
            consumo_diario_b = consumo_b / dias_facturados
            consumo_diario_c = consumo_c / dias_facturados
            consumo_diario_d = consumo_d / dias_facturados
            
            consumo_diario_total = consumo_diario_a + consumo_diario_b + consumo_diario_c + consumo_diario_d
            
            horas_pico_solares = 5
            
            demanda_consumo = consumo_diario_a + consumo_diario_b
            demanda_requerido = demanda_consumo / horas_pico_solares if horas_pico_solares > 0 else 0
            demanda_ideal = self.redondear_especial(demanda_requerido * 1.1)
            
            demanda24_consumo = consumo_diario_total
            demanda24_requerido = consumo_diario_total / horas_pico_solares if horas_pico_solares > 0 else 0
            demanda24_ideal = self.redondear_especial(demanda24_requerido * 1.1)
            
            respaldo24_consumo = demanda_consumo
            respaldo24_requerido = demanda_consumo
            respaldo24_ideal = self.redondear_especial(respaldo24_requerido * 1.1)
            
            respaldo4_consumo = consumo_diario_total
            respaldo4_requerido = consumo_diario_total / 4 if 4 > 0 else 0
            respaldo4_ideal = self.redondear_especial(respaldo4_requerido * 1.1)
            
            self.session_state['requerimiento_data'] = {
                "Demanda (08h00-22h00)_consumo": self.formatear_decimal(demanda_consumo),
                "Demanda (08h00-22h00)_requerido": self.formatear_decimal(demanda_requerido),
                "Demanda (08h00-22h00)_ideal": self.formatear_decimal(demanda_ideal),
                "Demanda (24 Horas)_consumo": self.formatear_decimal(demanda24_consumo),
                "Demanda (24 Horas)_requerido": self.formatear_decimal(demanda24_requerido),
                "Demanda (24 Horas)_ideal": self.formatear_decimal(demanda24_ideal),
                "Respaldo (24H)_consumo": self.formatear_decimal(respaldo24_consumo),
                "Respaldo (24H)_requerido": self.formatear_decimal(respaldo24_requerido),
                "Respaldo (24H)_ideal": self.formatear_decimal(respaldo24_ideal),
                "Respaldo (4 Horas de Respaldo)_consumo": self.formatear_decimal(respaldo4_consumo),
                "Respaldo (4 Horas de Respaldo)_requerido": self.formatear_decimal(respaldo4_requerido),
                "Respaldo (4 Horas de Respaldo)_ideal": self.formatear_decimal(respaldo4_ideal)
            }
            
        except Exception as e:
            st.error(f"Error al calcular requerimientos: {str(e)}")
            self.session_state['requerimiento_data'] = {
                k: "0.00" for k in self.session_state['requerimiento_data'].keys()
            }
    
    def mostrar_espacio_fisico(self):
        st.header("Espacio Físico Disponible")
        
        st.markdown("""
        <style>
        .array-table th {
            font-size: 12px;
            font-weight: bold;
            padding: 5px;
            text-align: center;
        }
        .array-table td {
            font-size: 12px;
            padding: 5px;
        }
        </style>
        """, unsafe_allow_html=True)
        
        try:
            num_arrays = int(float(self.session_state['cliente_data'].get("Arrays", "0"))) or 0
            num_arrays = max(0, min(num_arrays, 8))
        except:
            num_arrays = 0
        
        st.write(f"**Número de Arrays configurados:** {num_arrays}")

        cols = st.columns([2, 2, 2, 2])
        headers = ["**Array**", "**Medida X (m)**", "**Medida Y (m)**", "**Área (m²)**"]
        for i, header in enumerate(headers):
            cols[i].markdown(header, unsafe_allow_html=True)

        total_x = 0.0
        total_y = 0.0
        total_area = 0.0

        for i in range(1, num_arrays + 1):
            cols = st.columns([2, 2, 2, 2])
            
            cols[0].write(f"Array {i}")
            
            x_key = f"Array_{i}_X"
            x_val = self.session_state['array_data'].get(x_key, "0")
            new_x = cols[1].text_input(
                f"X_{i}",
                value=x_val,
                key=f"x_input_{i}",
                label_visibility="collapsed"
            )
            if new_x != x_val:
                self.session_state['array_data'][x_key] = new_x if new_x else "0"
                st.rerun()
            
            y_key = f"Array_{i}_Y"
            y_val = self.session_state['array_data'].get(y_key, "0")
            new_y = cols[2].text_input(
                f"Y_{i}",
                value=y_val,
                key=f"y_input_{i}",
                label_visibility="collapsed"
            )
            if new_y != y_val:
                self.session_state['array_data'][y_key] = new_y if new_y else "0"
                st.rerun()
            
            try:
                x = float(new_x.replace(",", ".")) if new_x and new_x.replace(",", "").replace(".", "").isdigit() else 0.0
                y = float(new_y.replace(",", ".")) if new_y and new_y.replace(",", "").replace(".", "").isdigit() else 0.0
                area = x * y
                total_x += x
                total_y += y
                total_area += area
                cols[3].write(f"{self.formatear_decimal(area)}")
            except:
                cols[3].write("0.00")
        
        if num_arrays > 0:
            st.subheader("Totales")
            cols = st.columns([2, 2, 2, 2])
            cols[0].write("**TOTAL**")
            cols[1].write(f"**{self.formatear_decimal(total_x)}**")
            cols[2].write(f"**{self.formatear_decimal(total_y)}**")
            cols[3].write(f"**{self.formatear_decimal(total_area)} m²**")
            
            if self.session_state['distribucion_data'].get('distribucion_actualizada', False):
                self.session_state['distribucion_data']['distribucion_actualizada'] = False
                st.rerun()
    
    def mostrar_requerimiento_energetico(self):
        self.calcular_requerimientos()
        
        st.header("Requerimiento Energético Solar y de Respaldo")
        
        st.markdown("""
        <style>
        .req-table {
            width: 100%;
        }
        .req-header {
            font-weight: bold !important;
            text-align: center !important;
            padding: 8px !important;
            background-color: #f0f2f6;
        }
        .req-input {
            width: 90% !important;
            padding: 6px !important;
            margin: 2px auto !important;
            font-size: 14px !important;
        }
        .req-label {
            font-size: 14px !important;
            padding: 8px !important;
            text-align: left !important;
        }
        </style>
        """, unsafe_allow_html=True)

        st.subheader("Demandas")
        col1, col2, col3, col4 = st.columns([3, 2, 2, 2])

        col1.markdown('<div class="req-header">Demandas</div>', unsafe_allow_html=True)
        col2.markdown('<div class="req-header">Consumo (kWh)</div>', unsafe_allow_html=True)
        col3.markdown('<div class="req-header">Requerido (kW)</div>', unsafe_allow_html=True)
        col4.markdown('<div class="req-header">Ideal (kW)</div>', unsafe_allow_html=True)

        demandas_completas = [
            "Demanda (08h00-22h00)",
            "Demanda (24 Horas)"
        ]

        for demanda in demandas_completas:
            cols = st.columns([3, 2, 2, 2])
            cols[0].markdown(f'<div class="req-label">{demanda}</div>', unsafe_allow_html=True)
            
            consumo_key = f"{demanda}_consumo"
            cols[1].text_input(
                f"Consumo {demanda}",
                value=self.session_state['requerimiento_data'].get(consumo_key, "0.00"),
                key=f"input_{consumo_key}",
                label_visibility="collapsed",
                disabled=True
            )
            
            req_key = f"{demanda}_requerido"
            new_req = cols[2].text_input(
                f"Requerido {demanda}",
                value=self.session_state['requerimiento_data'].get(req_key, "0.00"),
                key=f"input_{req_key}_editable",
                label_visibility="collapsed"
            )
            
            ideal_key = f"{demanda}_ideal"
            current_ideal = self.session_state['requerimiento_data'].get(ideal_key, "0.00")
            
            if new_req != self.session_state['requerimiento_data'].get(req_key, "0.00"):
                try:
                    req_value = float(new_req.replace(",", ".")) if new_req.replace(",", "").replace(".", "").isdigit() else 0.0
                    ideal_value = req_value * 1.1
                    ideal_value = self.redondear_especial(ideal_value)
                    
                    self.session_state['requerimiento_data'][req_key] = self.formatear_decimal(req_value)
                    self.session_state['requerimiento_data'][ideal_key] = self.formatear_decimal(ideal_value)
                    st.rerun()
                except:
                    pass
            
            new_ideal = cols[3].text_input(
                f"Ideal {demanda}",
                value=current_ideal,
                key=f"input_{ideal_key}_editable",
                label_visibility="collapsed"
            )
            
            if new_ideal != current_ideal:
                self.session_state['requerimiento_data'][ideal_key] = new_ideal
                st.rerun()

        st.subheader("Respaldo")
        col1, col2, col3 = st.columns([3, 2, 2])

        col1.markdown('<div class="req-header">Respaldo</div>', unsafe_allow_html=True)
        col2.markdown('<div class="req-header">Requerido (kW)</div>', unsafe_allow_html=True)
        col3.markdown('<div class="req-header">Ideal (kW)</div>', unsafe_allow_html=True)

        respaldos = [
            "Respaldo (24H)",
            "Respaldo (4 Horas de Respaldo)"
        ]

        for respaldo in respaldos:
            cols = st.columns([3, 2, 2])
            cols[0].markdown(f'<div class="req-label">{respaldo}</div>', unsafe_allow_html=True)
            
            req_key = f"{respaldo}_requerido"
            new_req = cols[1].text_input(
                f"Requerido {respaldo}",
                value=self.session_state['requerimiento_data'].get(req_key, "0.00"),
                key=f"input_{req_key}_editable_respaldo",
                label_visibility="collapsed"
            )
            
            ideal_key = f"{respaldo}_ideal"
            current_ideal = self.session_state['requerimiento_data'].get(ideal_key, "0.00")
            
            if new_req != self.session_state['requerimiento_data'].get(req_key, "0.00"):
                try:
                    req_value = float(new_req.replace(",", ".")) if new_req.replace(",", "").replace(".", "").isdigit() else 0.0
                    ideal_value = req_value * 1.1
                    ideal_value = self.redondear_especial(ideal_value)
                    
                    self.session_state['requerimiento_data'][req_key] = self.formatear_decimal(req_value)
                    self.session_state['requerimiento_data'][ideal_key] = self.formatear_decimal(ideal_value)
                    st.rerun()
                except:
                    pass
            
            new_ideal = cols[2].text_input(
                f"Ideal {respaldo}",
                value=current_ideal,
                key=f"input_{ideal_key}_editable_respaldo",
                label_visibility="collapsed"
            )
            
            if new_ideal != current_ideal:
                self.session_state['requerimiento_data'][ideal_key] = new_ideal
                st.rerun()

    
    def mostrar_busqueda_clientes(self):
        st.subheader("Buscar Cliente")

        desktop_path = Path.home() / 'Desktop'
        archivo = desktop_path / 'Calculadora_Solar' / 'Calculadora Solar.xlsx'

        if not archivo.exists():
            st.warning("⚠️ Archivo Excel no encontrado.")
            return

        try:
            self.df_clientes = pd.read_excel(archivo, sheet_name="Clientes")
        except Exception as e:
            st.error(f"❌ No se pudo leer la hoja 'Clientes': {e}")
            return

        # Asegurar columnas
        if "Empresa" not in self.df_clientes.columns:
            st.warning("⚠️ La hoja 'Clientes' no contiene la columna 'Empresa'.")
            return

        busqueda = st.text_input("Buscar por nombre o contacto:")

        if busqueda:
            filtrados = self.df_clientes[
                self.df_clientes["Empresa"].astype(str).str.contains(busqueda, case=False, na=False) |
                self.df_clientes.get("Persona de Contacto", "").astype(str).str.contains(busqueda, case=False, na=False)
            ]
        else:
            filtrados = self.df_clientes

        if filtrados.empty:
            st.warning("No se encontraron coincidencias.")
            return

        seleccion = st.selectbox(
            "Seleccione un cliente:",
            options=filtrados.index.tolist(),
            format_func=lambda idx: filtrados.loc[idx, "Empresa"] if pd.notna(filtrados.loc[idx, "Empresa"]) else f"Cliente #{idx}"
        )

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Cargar Cliente", use_container_width=True):
                self.cargar_cliente_seleccionado(seleccion)
                self.session_state['mostrar_busqueda'] = False
                st.rerun()

        with col2:
            if st.button("Cancelar", use_container_width=True):
                self.session_state['mostrar_busqueda'] = False
                st.rerun()
    
    def cargar_cliente_seleccionado(self, idx):
        from openpyxl import load_workbook
        import pandas as pd

        desktop_path = Path.home() / 'Desktop'
        archivo = desktop_path / 'Calculadora_Solar' / 'Calculadora Solar.xlsx'

        if not archivo.exists():
            st.error("❌ Archivo no encontrado.")
            return

        try:
            self.limpiar_campos() 
            empresa = self.df_clientes.loc[idx, "Empresa"]

            hojas = {
                "Clientes": "cliente_data",
                "Equipamiento Seleccionado": "equipamiento_seleccionado",
                "Distribución Solar": "distribucion_data",
                "Estructura Solar": "estructura_total_materiales",
                "Generación": "generacion_data",
                "Componentes Principales": "componentes_principales",
                "Costos": "costos_data",
                "Análisis Económico": "analisis_economico"
            }

            for hoja, key in hojas.items():
                try:
                    df = pd.read_excel(archivo, sheet_name=hoja)
                    if "Empresa" not in df.columns:
                       continue

                    df_cliente = df[df["Empresa"] == empresa]

                    if df_cliente.empty:
                        continue

                    if hoja == "Clientes":
                        # ✅ Usa método central que carga todo correctamente
                        self.cargar_datos_cliente(df_cliente.iloc[0])

                    elif hoja == "Estructura Solar":
                        materiales = df_cliente.iloc[0].get("Materiales", "{}")
                        import ast
                        self.session_state[key] = ast.literal_eval(materiales) if isinstance(materiales, str) else materiales

                    elif hoja == "Distribución Solar":
                        self.session_state[key] = {
                            "arrays_config": df_cliente.drop(columns=["Empresa"]).to_dict(orient="records")
                        }

                    elif hoja == "Costos":
                        self.session_state[key] = df_cliente.drop(columns=["Empresa"]).to_dict(orient="records")

                    else:
                        self.session_state[key] = df_cliente.iloc[0].drop(labels=["Empresa"]).to_dict()

                except Exception as e:
                    st.warning(f"⚠️ Error al cargar hoja '{hoja}': {e}")

            st.success(f"✅ Cliente '{empresa}' cargado correctamente.")
            st.experimental_rerun()

        except Exception as e:
            st.error(f"❌ Error al cargar cliente: {str(e)}")

    
    def actualizar_campos_tarifa(self, tarifa):
        if tarifa in self.tarifa_info:
            info = self.tarifa_info[tarifa]
            self.session_state['cliente_data']["Tarifario"] = info.get('tarifario', '')
            self.session_state['cliente_data']["Tensión eléctrica"] = info.get('tension', '')
            self.session_state['cliente_data']["Voltaje de entrada"] = info.get('voltaje', '')
    
    def limpiar_campos(self):
        self.session_state['cliente_data'] = {campo: "" for campo in self.campos_principales}
        self.session_state['consumo_data'] = {campo: "0" for campo in self.campos_consumo}
        self.session_state['array_data'] = {f"Array_{i}_{dim}": "0" for i in range(1, 9) for dim in ['X', 'Y']}
        self.session_state['requerimiento_data'] = {
            "Demanda (08h00-22h00)_consumo": "0.00",
            "Demanda (08h00-22h00)_requerido": "0.00",
            "Demanda (08h00-22h00)_ideal": "0.00",
            "Demanda (24 Horas)_consumo": "0.00",
            "Demanda (24 Horas)_requerido": "0.00",
            "Demanda (24 Horas)_ideal": "0.00",
            "Respaldo (24H)_consumo": "0.00",
            "Respaldo (24H)_requerido": "0.00",
            "Respaldo (24H)_ideal": "0.00",
            "Respaldo (4 Horas de Respaldo)_consumo": "0.00",
            "Respaldo (4 Horas de Respaldo)_requerido": "0.00",
            "Respaldo (4 Horas de Respaldo)_ideal": "0.00"
        }
        self.session_state['distribucion_data'] = {
            'panel_seleccionado': None,
            'paneles_por_array': {},
            'total_paneles': 0,
            'total_potencia': 0.0,
            'distribucion_actualizada': False,
            'arrays_config': [],
            'layout_data': {
                'orientacion': 'Portrait',
                'azimuth': '152°',
                'pitch': '5°',
                'ground_clearence': '',
                'espacio_bordes': '1,00 Metros',
                'filas_x_columna': '',
                'columnas': '',
                'total_filas_array': '',
                'paneles_fila': '',
                'paneles_columna': '',
                'paneles_array': '',
                'largo_fila': '',
                'separacion_columnas': '',
                'separacion_filas': ''
            }
        }
        self.session_state['equipamiento_seleccionado'] = {
            "Paneles Solares": "",
            "Inversores": "",
            "Baterías": "",
            "Convertidor de Alto Voltaje DC": "",
            "Materiales Estructura Solar": "",
            "Materiales DB": "",
            "Panel Solar": {},
            "Inversor": {}
        }
        self.session_state['estructura_total_materiales'] = []
        self.session_state['generacion_data'] = {}
        self.session_state['componentes_principales'] = {}
        self.session_state['costos_data'] = []
        self.session_state['analisis_economico'] = {}
        self.nuevo_cliente = True